import os
import queue
import sys
import threading
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox
from decimal import Decimal

import customtkinter as ctk
import tkinter.ttk as ttk

import nfe_parser
from nfe_parser import TRIBUTOS

try:
    from PIL import Image
    _TEM_PIL = True
except Exception:
    _TEM_PIL = False

try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass


def _base_dir():
    if getattr(sys, "frozen", False):
        if hasattr(sys, "_MEIPASS"):
            return sys._MEIPASS
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = _base_dir()
ICONE_ICO = os.path.join(BASE_DIR, "xml.ico")

COR_FUNDO = "#1a1c20"
COR_FUNDO_CARD = "#242424"
COR_BORDA_CARD = "#33363b"
COR_ACCENT = "#2fa460"
COR_TEXTO = "#e8eaed"
COR_TEXTO_FRACO = "#9aa0a6"
COR_VERMELHO = "#ef5350"
COR_AMARELO = "#ffb74d"

STATUS_CORES = {
    nfe_parser.STATUS_AUTORIZADA: "#4caf50",
    nfe_parser.STATUS_REJEITADA: "#ffa726",
    nfe_parser.STATUS_DENEGADA: "#ffb74d",
    nfe_parser.STATUS_CANCELADA: "#ef5350",
    nfe_parser.STATUS_SEM_PROTOCOLO: "#90a4ae",
    nfe_parser.STATUS_INUTILIZADA: "#ce93d8",
    nfe_parser.STATUS_CANCELAMENTO: "#ef5350",
    nfe_parser.STATUS_CCE: "#4fc3f7",
    nfe_parser.STATUS_EVENTO: "#80cbc4",
    nfe_parser.STATUS_INVALIDA: "#ef5350",
}


def fmt_brl(valor):
    if valor is None:
        return ""
    try:
        s = f"{Decimal(valor):,.2f}"
    except Exception:
        return str(valor)
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_emissao(data):
    if not data:
        return ""
    return data[:16].replace("T", " ")


def fmt_brl_total(valor):
    return "R$ " + fmt_brl(valor)


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("green")

        self.title("Calculadora de NF-e / NFC-e")
        self.geometry("980x680")
        self.minsize(820, 560)

        if os.path.exists(ICONE_ICO):
            try:
                self.iconbitmap(ICONE_ICO)
            except Exception:
                pass

        self._resumo = None
        self._analisando = False
        self._fila_progresso = queue.Queue()
        self._chips = []
        self._chips_trib = []
        self._apuracao_expanded = True

        self._font_titulo = ctk.CTkFont(family="Segoe UI", size=18, weight="bold")
        self._font_sub = ctk.CTkFont(family="Segoe UI", size=10)
        self._font_rotulo = ctk.CTkFont(family="Segoe UI", size=11)
        self._font_botao = ctk.CTkFont(family="Segoe UI", size=12, weight="bold")
        self._font_cartao_valor = ctk.CTkFont(family="Segoe UI", size=18, weight="bold")
        self._font_cartao_rotulo = ctk.CTkFont(family="Segoe UI", size=9)
        self._font_chip = ctk.CTkFont(family="Segoe UI", size=10)
        self._font_chip_valor = ctk.CTkFont(family="Segoe UI", size=10, weight="bold")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        self._criar_cabecalho()
        self._criar_card_pasta()
        self._criar_abas()
        self._criar_rodape()

    # ------------------------------------------------------------------ #
    # Cabeçalho
    # ------------------------------------------------------------------ #
    def _criar_cabecalho(self):
        cab = ctk.CTkFrame(self, corner_radius=0, fg_color="#16181c", height=58)
        cab.grid(row=0, column=0, sticky="ew")
        cab.grid_propagate(False)
        cab.grid_columnconfigure(1, weight=1)

        if os.path.exists(ICONE_ICO) and _TEM_PIL:
            try:
                img = ctk.CTkImage(dark_image=Image.open(ICONE_ICO), size=(36, 36))
                ctk.CTkLabel(cab, image=img, text="").grid(row=0, column=0,
                                                            padx=(16, 10), pady=10)
            except Exception:
                pass

        texto = ctk.CTkFrame(cab, fg_color="transparent")
        texto.grid(row=0, column=1, sticky="w", pady=8)
        ctk.CTkLabel(texto, text="Calculadora de NF-e / NFC-e",
                     font=self._font_titulo, text_color=COR_TEXTO).pack(anchor="w")
        ctk.CTkLabel(texto,
                     text="Apura impostos (ICMS, IPI, PIS, COFINS, FCP, IBS/CBS) de todos os XMLs da pasta",
                     font=self._font_sub, text_color=COR_TEXTO_FRACO).pack(anchor="w")

    # ------------------------------------------------------------------ #
    # Card de pasta
    # ------------------------------------------------------------------ #
    def _criar_card_pasta(self):
        card = ctk.CTkFrame(self, fg_color=COR_FUNDO_CARD, corner_radius=12,
                            border_width=1, border_color=COR_BORDA_CARD)
        card.grid(row=1, column=0, sticky="ew", padx=16, pady=(10, 4))
        card.grid_columnconfigure(1, weight=1)

        self.entry_pasta = ctk.CTkEntry(card, height=34,
                                        placeholder_text="Selecione a pasta onde estão os arquivos .xml",
                                        font=self._font_rotulo, fg_color="#2a2d31",
                                        border_color=COR_BORDA_CARD, text_color=COR_TEXTO)
        self.entry_pasta.grid(row=0, column=0, columnspan=3, sticky="ew", padx=14, pady=(12, 8))
        self.entry_pasta.bind("<Return>", lambda e: self.analisar())

        botoes = ctk.CTkFrame(card, fg_color="transparent")
        botoes.grid(row=1, column=0, columnspan=3, sticky="ew", padx=14, pady=(0, 4))
        botoes.grid_columnconfigure(0, weight=1)

        self.btn_selecionar = ctk.CTkButton(botoes, text="Selecionar Pasta", height=30,
                                            font=self._font_botao, corner_radius=8,
                                            command=self.selecionar_pasta)
        self.btn_selecionar.grid(row=0, column=0, sticky="w")

        self.btn_analisar = ctk.CTkButton(botoes, text="Analisar", height=30,
                                          font=self._font_botao, corner_radius=8,
                                          fg_color=COR_ACCENT, hover_color="#23894e",
                                          command=self.analisar)
        self.btn_analisar.grid(row=0, column=1, padx=(8, 0))

        self.btn_limpar = ctk.CTkButton(botoes, text="Limpar", height=30,
                                        font=self._font_botao, corner_radius=8,
                                        fg_color="#33363b", hover_color="#45494f",
                                        command=self.limpar)
        self.btn_limpar.grid(row=0, column=2, padx=(8, 0))

        self.var_dedup = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(card, text="Ignorar XMLs duplicados (mesma chave de acesso)",
                        variable=self.var_dedup, font=self._font_sub,
                        text_color=COR_TEXTO_FRACO, fg_color=COR_ACCENT,
                        hover_color="#23894e").grid(row=2, column=0, columnspan=3,
                                                    sticky="w", padx=14, pady=(2, 2))

        self.var_fcp = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(card, text="Considerar FCP (Fundo de Combate a Pobreza) no total",
                        variable=self.var_fcp, font=self._font_sub,
                        text_color=COR_TEXTO_FRACO, fg_color=COR_ACCENT,
                        hover_color="#23894e").grid(row=3, column=0, columnspan=3,
                                                    sticky="w", padx=14, pady=(2, 10))

    # ------------------------------------------------------------------ #
    # Abas
    # ------------------------------------------------------------------ #
    def _criar_abas(self):
        self.tabview = ctk.CTkTabview(self, corner_radius=10, fg_color=COR_FUNDO_CARD,
                                      border_width=1, border_color=COR_BORDA_CARD)
        self.tabview.grid(row=2, column=0, sticky="nsew", padx=16, pady=(4, 4))
        self.tabview.add("Resumo")
        self.tabview.add("Documentos")

        self._criar_aba_resumo(self.tabview.tab("Resumo"))
        self._criar_aba_documentos(self.tabview.tab("Documentos"))

    # ----------------------------- Resumo ----------------------------- #
    def _criar_aba_resumo(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)

        scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent", corner_radius=0)
        scroll.grid(row=0, column=0, sticky="nsew")
        scroll.grid_columnconfigure(0, weight=1)

        self._criar_cartoes(scroll)
        self._criar_resumo_status(scroll)
        self._criar_apuracao(scroll)

    def _criar_cartao(self, parent, col, titulo, valor_inicial, cor_valor=COR_TEXTO):
        cartao = ctk.CTkFrame(parent, fg_color=COR_FUNDO_CARD, corner_radius=10,
                              border_width=1, border_color=COR_BORDA_CARD)
        cartao.grid(row=0, column=col, sticky="nsew", padx=(5 if col else 0, 5))
        ctk.CTkLabel(cartao, text=titulo, font=self._font_cartao_rotulo,
                     text_color=COR_TEXTO_FRACO).pack(anchor="w", padx=12, pady=(8, 0))
        lbl = ctk.CTkLabel(cartao, text=valor_inicial, font=self._font_cartao_valor,
                           text_color=cor_valor)
        lbl.pack(anchor="w", padx=12, pady=(0, 8))
        return lbl

    def _criar_cartoes(self, parent):
        card = ctk.CTkFrame(parent, fg_color="transparent")
        card.grid(row=0, column=0, sticky="ew", padx=2, pady=(4, 6))
        for c in range(6):
            card.grid_columnconfigure(c, weight=1)

        self.lbl_total = self._criar_cartao(card, 0, "TOTAL AUTORIZADAS", "R$ 0,00", COR_ACCENT)
        self.lbl_qtd = self._criar_cartao(card, 1, "DOCUMENTOS", "0")
        self.lbl_qtd_nfe = self._criar_cartao(card, 2, "NF-e", "0")
        self.lbl_qtd_nfce = self._criar_cartao(card, 3, "NFC-e", "0")
        self.lbl_qtd_canceladas = self._criar_cartao(card, 4, "CANCELADAS", "0", COR_VERMELHO)
        self.lbl_total_trib = self._criar_cartao(card, 5, "TOTAL TRIBUTOS", "R$ 0,00", COR_AMARELO)

    def _criar_resumo_status(self, parent):
        card = ctk.CTkFrame(parent, fg_color=COR_FUNDO_CARD, corner_radius=12,
                            border_width=1, border_color=COR_BORDA_CARD)
        card.grid(row=1, column=0, sticky="ew", padx=2, pady=(4, 6))
        card.grid_columnconfigure(0, weight=1)
        self.status_card = card

        ctk.CTkLabel(card, text="Resumo por situação", font=self._font_cartao_rotulo,
                     text_color=COR_TEXTO_FRACO).grid(row=0, column=0, columnspan=5,
                                                       sticky="w", padx=12, pady=(8, 2))
        self.status_aviso = ctk.CTkLabel(card, text="Aguardando análise...",
                                        font=self._font_chip_valor,
                                        text_color=COR_TEXTO_FRACO)
        self.status_aviso.grid(row=1, column=0, columnspan=5, sticky="w", padx=12, pady=6)

    def _criar_chip_status(self, status, dados):
        cor = STATUS_CORES.get(status, COR_TEXTO)
        chip = ctk.CTkFrame(self.status_card, fg_color="#202226", corner_radius=8,
                            border_width=1, border_color=COR_BORDA_CARD)
        chip.grid_columnconfigure(1, weight=1)

        dot = ctk.CTkLabel(chip, text="●", text_color=cor, font=self._font_chip)
        dot.grid(row=0, column=0, padx=(8, 2), pady=(6, 0), sticky="w")
        ctk.CTkLabel(chip, text=status, text_color=COR_TEXTO, font=self._font_chip).grid(
            row=0, column=1, padx=(0, 8), pady=(6, 0), sticky="w")

        valor = fmt_brl_total(dados["total"]) if dados["total"] else ""
        linha = str(dados["qtd"])
        if valor:
            linha += "  ·  " + valor
        ctk.CTkLabel(chip, text=linha, text_color=COR_TEXTO_FRACO,
                     font=self._font_chip_valor).grid(row=1, column=0, columnspan=2,
                                                       padx=8, pady=(0, 6), sticky="w")
        return chip

    def _popular_status(self, por_status):
        for w in self._chips:
            w.destroy()
        self._chips = []
        self.status_aviso.grid_forget()

        col = 0
        for st in nfe_parser.STATUS_ORDEM:
            p = por_status.get(st)
            if not p or p["qtd"] == 0:
                continue
            chip = self._criar_chip_status(st, p)
            chip.grid(row=1 + col // 5, column=col % 5, padx=4, pady=(0, 8), sticky="nsew")
            self._chips.append(chip)
            col += 1
        if not self._chips:
            self.status_aviso.grid(row=1, column=0, columnspan=5, sticky="w", padx=12, pady=6)

    def _criar_apuracao(self, parent):
        card = ctk.CTkFrame(parent, fg_color=COR_FUNDO_CARD, corner_radius=12,
                            border_width=1, border_color=COR_BORDA_CARD)
        card.grid(row=2, column=0, sticky="ew", padx=2, pady=(4, 6))
        card.grid_columnconfigure(0, weight=1)

        cab = ctk.CTkFrame(card, fg_color="transparent")
        cab.grid(row=0, column=0, sticky="ew", padx=8, pady=(4, 2))
        cab.grid_columnconfigure(0, weight=1)
        self.apuracao_titulo = ctk.CTkLabel(cab, text="▾ Apuração de impostos (soma de todos os documentos)",
                                            font=self._font_cartao_rotulo, cursor="hand2",
                                            text_color=COR_TEXTO)
        self.apuracao_titulo.grid(row=0, column=0, sticky="w", padx=6)
        self.apuracao_titulo.bind("<Button-1>", lambda e: self._toggle_apuracao())

        self.apuracao_conteudo = ctk.CTkFrame(card, fg_color="transparent")
        self.apuracao_conteudo.grid(row=1, column=0, sticky="ew", padx=4, pady=(0, 6))
        self.apuracao_conteudo.grid_columnconfigure(0, weight=1)

        self.apuracao_aviso = ctk.CTkLabel(self.apuracao_conteudo, text="Aguardando análise...",
                                           font=self._font_chip_valor, text_color=COR_TEXTO_FRACO)
        self.apuracao_aviso.grid(row=0, column=0, sticky="w", padx=12, pady=6)

        self.apuracao_frame = ctk.CTkFrame(self.apuracao_conteudo, fg_color="transparent")
        self.apuracao_frame.grid(row=0, column=0, sticky="ew", padx=8, pady=(0, 4))
        self.apuracao_frame.grid_remove()
        for c in range(4):
            self.apuracao_frame.grid_columnconfigure(c, weight=1)

    def _toggle_apuracao(self):
        self._apuracao_expanded = not self._apuracao_expanded
        if self._apuracao_expanded:
            self.apuracao_titulo.configure(text="▾ Apuração de impostos (soma de todos os documentos)")
            self.apuracao_conteudo.grid(row=1, column=0, sticky="ew", padx=4, pady=(0, 6))
        else:
            self.apuracao_titulo.configure(text="▸ Apuração de impostos (soma de todos os documentos)")
            self.apuracao_conteudo.grid_remove()
        self.update_idletasks()

    def _popular_apuracao(self, tributos):
        for w in self._chips_trib:
            w.destroy()
        self._chips_trib = []
        self.apuracao_aviso.grid_remove()
        self.apuracao_frame.grid(row=0, column=0, sticky="nsew", padx=8, pady=(0, 4))

        col = 0
        sep_inserido = False
        for chave, titulo, _ in TRIBUTOS:
            if not sep_inserido and chave.startswith(("vIBS", "vCBS", "vBCIBSCBS")):
                linha_atual = col // 4
                sep = ctk.CTkLabel(self.apuracao_frame, text="Reforma Tributária (IBS / CBS)",
                                  font=self._font_rotulo, text_color=COR_ACCENT)
                sep.grid(row=linha_atual + 1, column=0, columnspan=4, sticky="w",
                         padx=4, pady=(8, 4))
                col = (linha_atual + 2) * 4
                sep_inserido = True
            val = tributos.get(chave, Decimal("0")) or Decimal("0")
            chip = ctk.CTkFrame(self.apuracao_frame, fg_color="#202226", corner_radius=8,
                                border_width=1, border_color=COR_BORDA_CARD)
            chip.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(chip, text=titulo, text_color=COR_TEXTO_FRACO,
                         font=self._font_chip_valor).grid(row=0, column=0, padx=8,
                                                          pady=(6, 0), sticky="w")
            ctk.CTkLabel(chip, text=fmt_brl_total(val), text_color=COR_TEXTO,
                         font=self._font_chip).grid(row=1, column=0, padx=8,
                                                    pady=(0, 6), sticky="w")
            chip.grid(row=col // 4, column=col % 4, padx=4, pady=4, sticky="nsew")
            self._chips_trib.append(chip)
            col += 1

    # --------------------------- Documentos --------------------------- #
    def _criar_aba_documentos(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)

        barra = ctk.CTkFrame(parent, fg_color="transparent")
        barra.grid(row=0, column=0, sticky="ew", padx=4, pady=(4, 4))
        barra.grid_columnconfigure(0, weight=1)

        self.var_mostrar_impostos = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(barra, text="Mostrar colunas de impostos na tabela (ICMS, IPI, PIS, IBS, CBS...)",
                        variable=self.var_mostrar_impostos, font=self._font_sub,
                        text_color=COR_TEXTO_FRACO, fg_color=COR_ACCENT,
                        hover_color="#23894e",
                        command=self._aplicar_visibilidade_colunas).grid(row=0, column=0, sticky="w")

        self.var_exportar_impostos = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(barra, text="Exportar impostos na planilha (desmarcar = formato antigo)",
                        variable=self.var_exportar_impostos, font=self._font_sub,
                        text_color=COR_TEXTO_FRACO, fg_color=COR_ACCENT,
                        hover_color="#23894e").grid(row=1, column=0, sticky="w", pady=(2, 0))

        self._criar_tabela(parent)

    def _criar_tabela(self, parent):
        card = ctk.CTkFrame(parent, fg_color=COR_FUNDO_CARD, corner_radius=12,
                            border_width=1, border_color=COR_BORDA_CARD)
        card.grid(row=1, column=0, sticky="nsew", padx=2, pady=(4, 4))
        card.grid_rowconfigure(1, weight=1)
        card.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(card, text="Documentos encontrados", font=self._font_rotulo,
                     text_color=COR_TEXTO).grid(row=0, column=0, sticky="w", padx=12, pady=(8, 4))

        BASE_COLS = ["modelo", "numero", "serie", "emissao", "emitente", "valor", "status", "arquivo"]
        TRIB_COLS = [chave for chave, _, _ in TRIBUTOS]
        self._base_cols = BASE_COLS
        self._todas_cols = BASE_COLS + TRIB_COLS

        self.tree = ttk.Treeview(card, columns=self._todas_cols, show="headings", height=10)
        colunas = {
            "modelo": ("Modelo", 58),
            "numero": ("Número", 62),
            "serie": ("Série", 48),
            "emissao": ("Emissão", 118),
            "emitente": ("Emitente", 170),
            "valor": ("Valor (R$)", 96),
            "status": ("Situação", 104),
            "arquivo": ("Arquivo", 150),
        }
        for chave, (titulo, larg) in colunas.items():
            self.tree.heading(chave, text=titulo)
            self.tree.column(chave, width=larg, anchor="w",
                             stretch=(chave in ("emitente", "arquivo")))
        self.tree.column("valor", anchor="e")
        for chave, titulo, larg in TRIBUTOS:
            self.tree.heading(chave, text=titulo)
            self.tree.column(chave, width=larg, anchor="e", stretch=False)
        self._aplicar_visibilidade_colunas()

        estilo = ttk.Style(self)
        try:
            estilo.theme_use("clam")
        except Exception:
            pass
        estilo.configure("Treeview", background="#202226", fieldbackground="#202226",
                         foreground=COR_TEXTO, borderwidth=0, rowheight=24,
                         font=("Segoe UI", 10))
        estilo.configure("Treeview.Heading", background="#2a2d31", foreground=COR_TEXTO,
                         borderwidth=0, relief="flat", font=("Segoe UI", 10, "bold"))
        estilo.map("Treeview.Heading", background=[("active", "#33363b")])
        estilo.map("Treeview", background=[("selected", "#1d4430")],
                   foreground=[("selected", COR_TEXTO)])

        frame_tab = ctk.CTkFrame(card, fg_color="#202226", corner_radius=6)
        frame_tab.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 8))
        frame_tab.grid_rowconfigure(0, weight=1)
        frame_tab.grid_columnconfigure(0, weight=1)

        self.tree.grid(row=0, column=0, sticky="nsew")

        self.scroll_y = ctk.CTkScrollbar(frame_tab, orientation="vertical",
                                         command=self.tree.yview)
        self.scroll_y.grid(row=0, column=1, sticky="ns")
        self.scroll_x = ctk.CTkScrollbar(frame_tab, orientation="horizontal",
                                         command=self.tree.xview)
        self.scroll_x.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=self.scroll_y.set,
                            xscrollcommand=self.scroll_x.set)

    def _aplicar_visibilidade_colunas(self):
        if not hasattr(self, "tree"):
            return
        if self.var_mostrar_impostos.get():
            self.tree["displaycolumns"] = self._todas_cols
        else:
            self.tree["displaycolumns"] = self._base_cols

    def _preencher_tabela(self, resumo):
        for item in self.tree.get_children():
            self.tree.delete(item)

        ordem = {st: i for i, st in enumerate(nfe_parser.STATUS_ORDEM)}

        def chave_status(doc):
            st = doc.status if not doc.erro else nfe_parser.STATUS_INVALIDA
            return ordem.get(st, 99)

        linhas = list(resumo["documentos"]) + list(resumo["extras"]) + list(resumo["erros"])
        linhas.sort(key=lambda d: (chave_status(d), os.path.basename(d.arquivo)))

        for doc in linhas:
            if doc.erro:
                status = nfe_parser.STATUS_INVALIDA
                emitente = doc.erro
            else:
                status = doc.status
                emitente = doc.emitente
            base = [doc.modelo_nome, doc.numero, doc.serie, fmt_emissao(doc.emissao),
                    emitente, fmt_brl(doc.valor), status, os.path.basename(doc.arquivo)]
            for chave, _, _ in TRIBUTOS:
                v = getattr(doc, chave, None)
                base.append(fmt_brl(v) if v is not None else "")
            self.tree.insert("", "end", tags=(status,), values=base)

        for status, cor in STATUS_CORES.items():
            self.tree.tag_configure(status, foreground=cor)

    # ------------------------------------------------------------------ #
    # Rodapé
    # ------------------------------------------------------------------ #
    def _criar_rodape(self):
        rodape = ctk.CTkFrame(self, fg_color="transparent")
        rodape.grid(row=3, column=0, sticky="ew", padx=16, pady=(2, 10))
        rodape.grid_columnconfigure(1, weight=1)

        self.lbl_status = ctk.CTkLabel(rodape, text="Selecione uma pasta para começar",
                                       font=self._font_sub, text_color=COR_TEXTO_FRACO)
        self.lbl_status.grid(row=0, column=0, sticky="w")

        self.progresso = ctk.CTkProgressBar(rodape, width=180, height=8,
                                            progress_color=COR_ACCENT)
        self.progresso.set(0)

        self.btn_exportar = ctk.CTkButton(rodape, text="Exportar XLSX", height=28,
                                          font=self._font_botao, corner_radius=8,
                                          fg_color="#33363b", hover_color="#45494f",
                                          command=self.exportar_xlsx, state="disabled")
        self.btn_exportar.grid(row=0, column=3, padx=(12, 0))

    # ------------------------------------------------------------------ #
    # Análise
    # ------------------------------------------------------------------ #
    def selecionar_pasta(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta com os XMLs de NF-e/NFC-e")
        if pasta:
            self.entry_pasta.delete(0, "end")
            self.entry_pasta.insert(0, pasta)
            self.analisar()

    def analisar(self):
        pasta = self.entry_pasta.get().strip().strip('"')
        if not pasta:
            messagebox.showinfo("Calculadora de NF-e/NFC-e", "Selecione uma pasta primeiro.")
            return
        if not os.path.isdir(pasta):
            messagebox.showwarning("Calculadora de NF-e/NFC-e",
                                   "A pasta informada não existe:\n" + pasta)
            return
        if self._analisando:
            return

        self._analisando = True
        self.btn_selecionar.configure(state="disabled")
        self.btn_analisar.configure(state="disabled")
        self.btn_limpar.configure(state="disabled")
        self.btn_exportar.configure(state="disabled")
        self.progresso.set(0)
        self.progresso.grid(row=0, column=1, padx=(12, 0))
        self.lbl_status.configure(text="Analisando arquivos...")

        threading.Thread(target=self._trabalho, args=(pasta,), daemon=True).start()
        self.after(80, self._poll_progresso)

    def _progresso(self, i, total, nome):
        self._fila_progresso.put((i, total, nome))

    def _poll_progresso(self):
        if not self._analisando:
            return
        try:
            while True:
                i, total, nome = self._fila_progresso.get_nowait()
                if total:
                    self.progresso.set(i / total)
                self.lbl_status.configure(text=f"Analisando ({i}/{total})... {nome}")
        except queue.Empty:
            pass
        self.after(80, self._poll_progresso)

    def _trabalho(self, pasta):
        try:
            docs = nfe_parser.escanear_pasta(pasta, incluir_subpastas=True,
                                             on_progress=self._progresso)
            resumo = nfe_parser.resumir(docs, ignorar_duplicados=self.var_dedup.get(),
                                         considerar_fcp=self.var_fcp.get())
            self.after(0, lambda: self._concluir(resumo))
        except Exception as e:
            self.after(0, lambda: self._falha(e))

    def _falha(self, e):
        self._analisando = False
        self._fim_estados()
        self.progresso.grid_forget()
        self.lbl_status.configure(text="Erro ao analisar a pasta.")
        messagebox.showerror("Calculadora de NF-e/NFC-e", f"Erro ao analisar:\n{e}")

    def _concluir(self, resumo):
        self._analisando = False
        self._resumo = resumo
        self._fim_estados()
        self.progresso.grid_forget()

        self.lbl_total.configure(text=fmt_brl_total(resumo["total"]))
        self.lbl_qtd.configure(text=str(resumo["quantidade"]))
        self.lbl_qtd_nfe.configure(text=str(resumo["qtd_nfe"]))
        self.lbl_qtd_nfce.configure(text=str(resumo["qtd_nfce"]))
        self.lbl_qtd_canceladas.configure(text=str(resumo["qtd_canceladas"]))
        self.lbl_total_trib.configure(text=fmt_brl_total(resumo["tributos"].get("vTotTrib", Decimal("0"))))

        self._popular_status(resumo["por_status"])
        self._popular_apuracao(resumo["tributos"])
        self._preencher_tabela(resumo)
        self.btn_exportar.configure(state="normal" if resumo["documentos"] else "disabled")

        partes = [
            f"{resumo['quantidade']} documento(s) somados",
            f"{resumo['qtd_canceladas']} cancelada(s)",
            f"{resumo['qtd_extras']} evento(s)/registro(s)",
            f"{resumo['qtd_ignoradas']} ignorado(s)",
        ]
        self.lbl_status.configure(text=" · ".join(partes))

    def _fim_estados(self):
        self.btn_selecionar.configure(state="normal")
        self.btn_analisar.configure(state="normal")
        self.btn_limpar.configure(state="normal")

    def limpar(self):
        if self._analisando:
            return
        self.entry_pasta.delete(0, "end")
        self._resumo = None
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.lbl_total.configure(text="R$ 0,00")
        self.lbl_qtd.configure(text="0")
        self.lbl_qtd_nfe.configure(text="0")
        self.lbl_qtd_nfce.configure(text="0")
        self.lbl_qtd_canceladas.configure(text="0")
        self.lbl_total_trib.configure(text="R$ 0,00")
        for w in self._chips:
            w.destroy()
        self._chips = []
        self.status_aviso.configure(text="Aguardando análise...")
        self.status_aviso.grid(row=1, column=0, columnspan=5, sticky="w", padx=12, pady=6)
        for w in self._chips_trib:
            w.destroy()
        self._chips_trib = []
        self.apuracao_frame.grid_remove()
        self.apuracao_aviso.grid(row=0, column=0, sticky="w", padx=12, pady=6)
        self._apuracao_expanded = True
        self.apuracao_titulo.configure(text="▾ Apuração de impostos (soma de todos os documentos)")
        self.apuracao_conteudo.grid(row=1, column=0, sticky="ew", padx=4, pady=(0, 6))
        self.lbl_status.configure(text="Selecione uma pasta para começar")
        self.btn_exportar.configure(state="disabled")

    # ------------------------------------------------------------------ #
    # Exportação XLSX
    # ------------------------------------------------------------------ #
    def exportar_xlsx(self):
        if not self._resumo or not self._resumo["documentos"]:
            messagebox.showinfo("Calculadora de NF-e/NFC-e", "Nada para exportar.")
            return
        caminho = filedialog.asksaveasfilename(
            title="Salvar resumo em Excel",
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx")],
            initialfile="resumo_nfes_nfces.xlsx")
        if not caminho:
            return
        try:
            from openpyxl import Workbook
            exportar_trib = self.var_exportar_impostos.get()
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumo"

            cabecalhos = ["Modelo", "Numero", "Serie", "Emissao", "Emitente", "Valor (R$)", "Situacao", "Chave"]
            if exportar_trib:
                cabecalhos += [titulo for _, titulo, _ in TRIBUTOS]

            ws.append(cabecalhos)

            todas = (list(self._resumo["documentos"])
                     + list(self._resumo["extras"])
                     + list(self._resumo["erros"]))
            for doc in todas:
                if doc.erro:
                    linha = ["", "", "", "", doc.erro, "", "Invalida", ""]
                    if exportar_trib:
                        linha += [0 for _ in TRIBUTOS]
                else:
                    linha = [doc.modelo_nome, doc.numero, doc.serie,
                             fmt_emissao(doc.emissao), doc.emitente,
                             float(doc.valor) if doc.valor is not None else 0,
                             doc.status, doc.chave]
                    if exportar_trib:
                        for chave, _, _ in TRIBUTOS:
                            v = getattr(doc, chave, None)
                            linha.append(float(v) if v is not None else 0)
                ws.append(linha)

            ws.append([])

            moeda = 'R$ #,##0.00'

            linha_resumo = ws.max_row + 1
            ws.cell(row=linha_resumo, column=1, value="RESUMO POR SITUACAO")
            ws.cell(row=linha_resumo, column=2, value="Quantidade")
            ws.cell(row=linha_resumo, column=3, value="Valor (R$)")
            for st in nfe_parser.STATUS_ORDEM:
                p = self._resumo["por_status"].get(st)
                if p and p["qtd"]:
                    linha_resumo += 1
                    ws.cell(row=linha_resumo, column=1, value=st)
                    ws.cell(row=linha_resumo, column=2, value=p["qtd"])
                    c = ws.cell(row=linha_resumo, column=3, value=float(p["total"]))
                    c.number_format = moeda

            if exportar_trib:
                linha_resumo += 2
                ws.cell(row=linha_resumo, column=1, value="APURACAO DE IMPOSTOS")
                ws.cell(row=linha_resumo, column=2, value="Valor (R$)")
                for chave, titulo, _ in TRIBUTOS:
                    linha_resumo += 1
                    ws.cell(row=linha_resumo, column=1, value=titulo)
                    val = self._resumo["tributos"].get(chave, Decimal("0")) or Decimal("0")
                    c = ws.cell(row=linha_resumo, column=2, value=float(val))
                    c.number_format = moeda

            linha_resumo += 2
            ws.cell(row=linha_resumo, column=1, value="TOTAL AUTORIZADAS")
            ws.cell(row=linha_resumo, column=2, value=self._resumo["quantidade"])
            c = ws.cell(row=linha_resumo, column=3, value=float(self._resumo["total"]))
            c.number_format = moeda

            if self.var_fcp.get():
                ws.cell(row=linha_resumo, column=5, value="FCP considerado no total")

            self._estilizar_xlsx(ws)

            wb.save(caminho)
            messagebox.showinfo("Calculadora de NF-e/NFC-e",
                                 "Planilha exportada com sucesso:\n" + caminho)
        except Exception as e:
            messagebox.showerror("Calculadora de NF-e/NFC-e",
                                  f"Erro ao exportar:\n{e}")

    def _estilizar_xlsx(self, ws):
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        moeda_cols = {"Valor (R$)"}
        moeda_cols.update(titulo for _, titulo, _ in TRIBUTOS)

        azul = PatternFill("solid", fgColor="1F4E78")
        cinza = PatternFill("solid", fgColor="F2F2F2")
        branco_font = Font(color="FFFFFF", bold=True, size=11)
        borda = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"), bottom=Side(style="thin"))

        for celula in ws[1]:
            celula.fill = azul
            celula.font = branco_font
            celula.alignment = Alignment(horizontal="center")
            celula.border = borda

        for linha in ws.iter_rows(min_row=2):
            for celula in linha:
                celula.border = borda
            if linha[0].row % 2 == 0:
                for celula in linha:
                    celula.fill = cinza

        for coluna in ws.columns:
            maior = 0
            letra = get_column_letter(coluna[0].column)
            for celula in coluna:
                if celula.value is not None:
                    tamanho = len(str(celula.value))
                    if tamanho > maior:
                        maior = tamanho
            ws.column_dimensions[letra].width = maior + 5

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for coluna in ws.iter_cols():
            titulo = str(coluna[0].value or "")
            if titulo in moeda_cols:
                for celula in coluna[1:]:
                    if isinstance(celula.value, (int, float)):
                        celula.number_format = 'R$ #,##0.00'


if __name__ == "__main__":
    App().mainloop()
